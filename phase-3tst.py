import pandas as pd
import numpy as np
import os
from datetime import time

# ======================================================
# CONFIG
# ======================================================

BASE_DIR = "downloaded_data"
NSEI_FILE = "downloaded_data/NSEI/intraday_5m.csv"
PHASE2_FILE = "phase-2results/Phase_2.xlsx"

MARKET_OPEN = time(9, 15)
ORB_END = time(9, 30)
MODE_A_START = time(9, 45)

TICK_SIZE = 0.05

# ======================================================
# PULSE915 PHASE 3 STOP-LOSS CONFIGURATION
# ======================================================

# ATR-based stop parameters
ATR_MULTIPLIER = 1.25  # k: ATR multiplier for volatility stop
STOP_MIN_PCT = 1.0     # sₘᵢₙ: Minimum stop distance percent
STOP_MAX_PCT = 2.5     # sₘₐₓ: Maximum stop distance percent

# Risk-Reward ratio
RISK_REWARD_RATIO = 2.0  # R: Target is 2x the risk (2:1 reward:risk)

# ATR calculation parameters
ATR_LENGTH_5M = 14  # ATR period for 5-minute candles

# ======================================================
# LOAD PHASE 2
# ======================================================

def load_phase2():
    """Load Phase 2 results - we only need Symbol and Date to know which stocks qualified"""
    df = pd.read_excel(PHASE2_FILE)
    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    # We'll compute RS_30m and VolMult_od dynamically, so just load Symbol and Date
    return df[["Symbol", "Date"]].dropna()

# ======================================================
# LOAD STOCK 5M
# ======================================================

def load_stock_5m(symbol):
    folder = os.path.join(BASE_DIR, symbol)
    dfs = []

    if not os.path.exists(folder):
        return None

    for file in sorted(os.listdir(folder)):
        if not file.endswith(".csv"):
            continue

        file_date = pd.to_datetime(file.replace(".csv", "")).date()
        df = pd.read_csv(os.path.join(folder, file))

        df["Datetime"] = pd.to_datetime(
            df["Time"].astype(str).apply(lambda t: f"{file_date} {t}"),
            utc=True,
            errors="coerce"
        )

        for c in ["Open", "High", "Low", "Close", "Volume"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")

        df = df[df["Datetime"].notna()]
        df["Date"] = file_date
        df["Symbol"] = symbol
        dfs.append(df)

    if not dfs:
        return None

    return pd.concat(dfs).sort_values("Datetime").reset_index(drop=True)

# ======================================================
# LOAD NSEI
# ======================================================

def load_nsei_5m():
    df = pd.read_csv(NSEI_FILE)
    df["Datetime"] = pd.to_datetime(df["Datetime"], utc=True)
    df["Close"] = pd.to_numeric(df["Close"], errors="coerce")
    return df[["Datetime", "Close"]].rename(columns={"Close": "NIFTY_Close"})

# ======================================================
# METRICS
# ======================================================

def compute_vwap(df):
    """Compute intraday VWAP that resets daily"""
    tp = (df["High"] + df["Low"] + df["Close"]) / 3
    df = df.copy()
    df["tp_vol"] = tp * df["Volume"]
    df["cumsum_tp_vol"] = df.groupby("Date")["tp_vol"].cumsum()
    df["cumsum_vol"] = df.groupby("Date")["Volume"].cumsum()
    return df["cumsum_tp_vol"] / df["cumsum_vol"]

def compute_daily_atr(df):
    daily = df.groupby("Date").agg(
        High=("High", "max"),
        Low=("Low", "min"),
        Close=("Close", "last")
    )

    prev_close = daily["Close"].shift(1)
    tr = pd.concat([
        daily["High"] - daily["Low"],
        (daily["High"] - prev_close).abs(),
        (daily["Low"] - prev_close).abs()
    ], axis=1).max(axis=1)

    # Use Wilder's EMA method for ATR (alpha = 1/14)
    atr14 = tr.ewm(alpha=1/14, min_periods=1, adjust=False).mean()
    daily["ATR_pct"] = (atr14 / daily["Close"]) * 100
    return daily[["ATR_pct"]]

def round_to_tick(price):
    return np.ceil(price / TICK_SIZE) * TICK_SIZE

def round_to_tick_down(price):
    """Round price DOWN to nearest tick (for stop-loss)"""
    return np.floor(price / TICK_SIZE) * TICK_SIZE

def round_to_tick_up(price):
    """Round price UP to nearest tick (for target)"""
    return np.ceil(price / TICK_SIZE) * TICK_SIZE

def compute_atr_5m(df, length=14):
    """
    Compute intraday ATR on 5-minute candles
    Returns ATR% = (ATR / Close) × 100
    """
    df = df.copy()
    df = df.sort_values(['Symbol', 'Date', 'Datetime'])
    
    # Calculate True Range for each candle
    df['prev_close'] = df.groupby(['Symbol', 'Date'])['Close'].shift(1)
    
    df['tr'] = pd.concat([
        df['High'] - df['Low'],
        (df['High'] - df['prev_close']).abs(),
        (df['Low'] - df['prev_close']).abs()
    ], axis=1).max(axis=1)
    
    # Calculate rolling ATR within each day using Wilder's EMA method
    # Wilder's smoothing: alpha = 1/length
    df['ATR_5m'] = df.groupby(['Symbol', 'Date'])['tr'].transform(
        lambda x: x.ewm(alpha=1/length, min_periods=1, adjust=False).mean()
    )
    
    # Convert to percentage
    df['ATR_5m_pct'] = (df['ATR_5m'] / df['Close']) * 100
    
    return df

def calculate_stop_loss_and_target(row, mode, orb_low=None, consolidation_low=None):
    """
    Calculate stop-loss and target for a confirmed entry
    
    Parameters:
    - row: DataFrame row with entry data
    - mode: 'A', 'B', or 'C'
    - orb_low: Opening Range Low (for Mode A)
    - consolidation_low: Consolidation Low (for Mode C)
    
    Returns:
    - dict with stop_atr, stop_structure, final_stop, target, risk_per_share, risk_pct
    """
    entry_price = row[f'Mode{mode}_Entry']
    
    if pd.isna(entry_price):
        return None
    
    # 1. Calculate micro-buffer δ
    delta = max(2 * TICK_SIZE, 0.0005 * entry_price)
    
    # 2. ATR-based stop
    atr_pct_5m = row['ATR_5m_pct'] if pd.notna(row.get('ATR_5m_pct')) else 1.5  # Default to 1.5% if not available or NaN
    
    # Clip ATR stop distance to bounds
    stop_distance_pct = np.clip(
        ATR_MULTIPLIER * atr_pct_5m,
        STOP_MIN_PCT,
        STOP_MAX_PCT
    ) / 100  # Convert to decimal
    
    delta_atr = entry_price * stop_distance_pct
    stop_atr = entry_price - delta_atr
    
    # 3. Structure-based stop (mode-dependent)
    vwap = row.get('VWAP', entry_price)
    
    if mode == 'A':
        # Mode A: ORB Breakout - stop below ORB Low
        if orb_low is not None:
            stop_structure = orb_low - delta
        else:
            # Fallback: use VWAP if ORB low not available
            stop_structure = min(orb_low if orb_low else vwap, vwap) - delta
    
    elif mode == 'B':
        # Mode B: VWAP Reclaim - stop below VWAP
        stop_structure = vwap - delta
    
    elif mode == 'C':
        # Mode C: Day-High Continuation - stop below consolidation low
        if consolidation_low is not None:
            stop_structure = consolidation_low - delta
        else:
            # Fallback: use recent low from last 6 candles
            stop_structure = row.get('Low', entry_price * 0.98) - delta
    
    else:
        stop_structure = stop_atr  # Fallback
    
    # 4. Final stop selection (hybrid rule)
    # Choose the tighter stop that respects both structure and volatility
    final_stop = max(stop_structure, stop_atr)
    
    # 5. Tick rounding
    final_stop = round_to_tick_down(final_stop)
    
    # 6. Calculate risk and target
    risk_per_share = entry_price - final_stop
    risk_pct = (risk_per_share / entry_price) * 100
    
    target = entry_price + (RISK_REWARD_RATIO * risk_per_share)
    target = round_to_tick_up(target)
    
    return {
        'Stop_ATR': round(stop_atr, 2),
        'Stop_Structure': round(stop_structure, 2),
        'Final_Stop': round(final_stop, 2),
        'Risk_Per_Share': round(risk_per_share, 2),
        'Risk_Pct': round(risk_pct, 2),
        'Target': round(target, 2)
    }

def compute_rs_30m(df):
    """
    Compute RS_30m dynamically for each 5m candle
    RS_30m(t) = R_stock_30m(t) - R_nifty_30m(t)
    where R = ((Close - Close_30m_ago) / Close_30m_ago) × 100
    30 minutes = 6 candles at 5m intervals
    """
    # Stock return over last 30 minutes (6 candles)
    df["R_stock_30m"] = ((df["Close"] - df["Close"].shift(6)) / df["Close"].shift(6)) * 100
    
    # NIFTY return over last 30 minutes (6 candles)
    df["R_nifty_30m"] = ((df["NIFTY_Close"] - df["NIFTY_Close"].shift(6)) / df["NIFTY_Close"].shift(6)) * 100
    
    # Relative strength
    df["RS_30m"] = df["R_stock_30m"] - df["R_nifty_30m"]
    
    return df

def compute_volmult_od(df):
    """
    Compute VolMult_od(t) using HISTORICAL time-of-day average (NO FUTURE LEAK)
    VolMult_od(t) = Vol_od(t) / Expected_Vol_od(t)
    where:
      Vol_od(t) = cumulative volume from 09:15 to time t TODAY
      Expected_Vol_od(t) = Historical avg volume at this time-of-day × candles_elapsed
      
    Since we don't have multi-day history per stock, we use expanding mean from
    PREVIOUS days only (excludes current day to avoid future leak)
    """
    # Add time-of-day marker (candle number within day)
    df["Candle_num"] = df.groupby("Date").cumcount() + 1
    
    # Cumulative volume from start of day
    df["Vol_od"] = df.groupby("Date")["Volume"].cumsum()
    
    # For each symbol, compute historical average volume per candle position
    # Using expanding mean EXCLUDING current day (shift by 1 day)
    df_sorted = df.sort_values(["Symbol", "Date", "Datetime"])
    
    # Group by symbol and candle number, then compute 5-day rolling mean excluding current row
    historical_avg = df_sorted.groupby(["Symbol", "Candle_num"])["Volume"].apply(
        lambda x: x.shift(1).rolling(window=5, min_periods=1).mean()
    ).reset_index(level=[0, 1], drop=True)
    
    df_sorted["Historical_Avg_Vol"] = historical_avg
    
    # For first occurrence (no history), use overall symbol average as fallback
    symbol_avg = df_sorted.groupby("Symbol")["Volume"].transform("mean")
    df_sorted["Historical_Avg_Vol"] = df_sorted["Historical_Avg_Vol"].fillna(symbol_avg)
    
    # Expected cumulative volume = historical_avg × candles_elapsed
    df_sorted["Expected_Vol_od"] = df_sorted["Historical_Avg_Vol"] * df_sorted["Candle_num"]
    
    # VolMult_od = actual / expected
    df_sorted["VolMult_od"] = df_sorted["Vol_od"] / df_sorted["Expected_Vol_od"]
    
    # Fill any NaN or inf with 1.0 (neutral)
    df_sorted["VolMult_od"] = df_sorted["VolMult_od"].replace([np.inf, -np.inf], 1.0).fillna(1.0)
    
    # Merge back to original df
    df = df.merge(
        df_sorted[["Symbol", "Date", "Datetime", "Vol_od", "Candle_num", "VolMult_od"]],
        on=["Symbol", "Date", "Datetime"],
        how="left",
        suffixes=("", "_new")
    )
    
    # Use the new VolMult_od
    if "VolMult_od_new" in df.columns:
        df["VolMult_od"] = df["VolMult_od_new"]
        df = df.drop(columns=["VolMult_od_new"])
    
    return df

# ======================================================
# PHASE 3
# ======================================================

def run_phase3_for_symbol(symbol, phase2_df, nsei_df):
    stock = load_stock_5m(symbol)
    if stock is None:
        return None

    p2 = phase2_df[phase2_df["Symbol"] == symbol]

    # Merge with NSEI data
    df = pd.merge(stock, nsei_df, on="Datetime", how="inner")
    
    # Only keep dates that passed Phase 2
    df = pd.merge(df, p2, on=["Symbol", "Date"], how="inner")
    
    if df.empty:
        return None

    # Compute dynamic metrics
    df["VWAP"] = compute_vwap(df)
    df = compute_rs_30m(df)  # Adds RS_30m, R_stock_30m, R_nifty_30m
    df = compute_volmult_od(df)  # Adds VolMult_od
    
    # Compute 5-minute ATR for stop-loss calculation
    df = compute_atr_5m(df, length=ATR_LENGTH_5M)  # Adds ATR_5m, ATR_5m_pct
    
    df["Time"] = df["Datetime"].dt.time

    atr_daily = compute_daily_atr(df)
    df = df.merge(atr_daily, on="Date", how="left")

    df["Buffer"] = df["Close"] * np.maximum(
        0.0005,
        0.10 * (df["ATR_pct"] / 100)
    )

    # ================= MODES =================
    for m in ["A", "B", "C"]:
        df[f"Mode{m}_Eligible"] = False
        df[f"Mode{m}_Trigger"] = np.nan
        df[f"Mode{m}_Confirmed"] = False
        df[f"Mode{m}_Entry"] = np.nan

    df["ORBHigh"] = np.nan
    df["ORBLow"] = np.nan  # Track ORB Low for Mode A stop-loss
    df["DayHigh"] = np.nan
    df["NearVWAP"] = abs((df["Close"] - df["VWAP"]) / df["VWAP"]) <= 0.0025
    df["NearHigh"] = False

    for day in df["Date"].unique():
        day_df = df[df["Date"] == day].copy()

        # ---------- MODE C ----------
        day_df["DayHigh"] = day_df["High"].cummax()
        df.loc[day_df.index, "DayHigh"] = day_df["DayHigh"]

        near_high = (day_df["DayHigh"] - day_df["Close"]) / day_df["DayHigh"] <= 0.004
        df.loc[day_df.index, "NearHigh"] = near_high

        # Mode C: Near high with volume
        eligible_c = near_high & (day_df["VolMult_od"] >= 1.5)
        eligible_c_idx = day_df[eligible_c].index
        
        if len(eligible_c_idx) > 0:
            df.loc[eligible_c_idx, "ModeC_Eligible"] = True
            # Calculate trigger using day_df values to ensure alignment
            df.loc[eligible_c_idx, "ModeC_Trigger"] = round_to_tick(
                day_df.loc[eligible_c_idx, "DayHigh"] + day_df.loc[eligible_c_idx, "Buffer"]
            )

        # ---------- MODE A ----------
        orb = day_df[(day_df["Time"] >= MARKET_OPEN) & (day_df["Time"] <= ORB_END)]
        if len(orb) >= 3:
            orb_high = orb["High"].max()
            orb_low = orb["Low"].min()
            df.loc[df["Date"] == day, "ORBHigh"] = orb_high
            df.loc[df["Date"] == day, "ORBLow"] = orb_low

            # Mode A: within first ~45 minutes (09:45 to 10:00)
            mode_a_end = time(10, 0)
            
            eligible_a = (
                (df["Date"] == day) &
                (df["Time"] >= MODE_A_START) &
                (df["Time"] <= mode_a_end) &  # FIX: Add upper time bound
                (df["Close"] > df["VWAP"]) &
                (df["VolMult_od"] >= 1.8) &
                (df["RS_30m"] >= 0.6)
            )

            df.loc[eligible_a, "ModeA_Eligible"] = True
            df.loc[eligible_a, "ModeA_Trigger"] = round_to_tick(
                orb_high + df.loc[eligible_a, "Buffer"]
            )

        # ---------- MODE B ----------
        prev_close = day_df["Close"].shift(1)
        prev_vwap = day_df["VWAP"].shift(1)
        reclaim = (prev_close <= prev_vwap) & (day_df["Close"] > day_df["VWAP"])

        eligible_b = reclaim & day_df["NearVWAP"] & (day_df["VolMult_od"] >= 1.3)
        eligible_b_idx = day_df[eligible_b].index
        
        if len(eligible_b_idx) > 0:
            df.loc[eligible_b_idx, "ModeB_Eligible"] = True
            # Calculate trigger using day_df values to ensure alignment
            df.loc[eligible_b_idx, "ModeB_Trigger"] = round_to_tick(
                day_df.loc[eligible_b_idx, "VWAP"] + day_df.loc[eligible_b_idx, "Buffer"]
            )
    # ================= MODE PRIORITY ENFORCEMENT =================
    # Priority: A > B > C (per Symbol per Date)
    # Only clear ELIGIBILITY flags, keep triggers for reference

    for (sym, date), g in df.groupby(["Symbol", "Date"]):

        has_A = g["ModeA_Eligible"].any()
        has_B = g["ModeB_Eligible"].any()

        if has_A:
            # Kill B & C eligibility (but keep their triggers for reference)
            df.loc[g.index, "ModeB_Eligible"] = False
            df.loc[g.index, "ModeC_Eligible"] = False

        elif has_B:
            # Kill C eligibility only (but keep trigger for reference)
            df.loc[g.index, "ModeC_Eligible"] = False

        # else → only Mode C survives (or nothing)

    # ================= PHASE 3C CONFIRMATION =================
    # CRITICAL: Only confirm modes that are STILL ELIGIBLE after priority enforcement
    # Wait for 5m candle close ABOVE trigger (Entry > Trigger, not Entry < Trigger!)
    
    for mode in ["A", "B", "C"]:
        elig = f"Mode{mode}_Eligible"
        trig = f"Mode{mode}_Trigger"
        conf = f"Mode{mode}_Confirmed"
        entry = f"Mode{mode}_Entry"

        for (sym, date), g in df.groupby(["Symbol", "Date"]):
            g = g.sort_values("Datetime")
            
            # CRITICAL: Check if this mode is STILL eligible for this symbol-date
            # (after mode priority enforcement)
            if not g[elig].any():
                continue  # Mode was killed by priority enforcement, skip
            
            # Find candles where trigger was set AND mode is still eligible
            eligible_with_trigger = g[(g[elig]) & (g[trig].notna())]
            
            if eligible_with_trigger.empty:
                continue
            
            # Get the trigger value from first eligible candle
            trigger_value = eligible_with_trigger[trig].iloc[0]
            trigger_time_idx = eligible_with_trigger.index[0]
            
            # Find first candle AFTER trigger set where Close > Trigger
            # CRITICAL: Entry must be GREATER than trigger, not less!
            subsequent_candles = g.loc[trigger_time_idx:]
            confirmed = subsequent_candles[subsequent_candles["Close"] > trigger_value]
            
            if not confirmed.empty:
                idx = confirmed.index[0]
                df.loc[idx, conf] = True
                df.loc[idx, entry] = df.loc[idx, "Close"]
                # Also save the trigger value on the confirmed candle for easy reference
                df.loc[idx, trig] = trigger_value

    # ================= STOP-LOSS AND TARGET CALCULATION =================
    # Calculate stop-loss and target for all confirmed entries
    # Using Pulse915 Phase 3 Stop-Loss methodology
    
    # Initialize columns for stop-loss and target data
    for mode in ["A", "B", "C"]:
        df[f"Mode{mode}_Stop_ATR"] = np.nan
        df[f"Mode{mode}_Stop_Structure"] = np.nan
        df[f"Mode{mode}_Final_Stop"] = np.nan
        df[f"Mode{mode}_Target"] = np.nan
        df[f"Mode{mode}_Risk_Per_Share"] = np.nan
        df[f"Mode{mode}_Risk_Pct"] = np.nan
    
    # Process each confirmed entry
    for mode in ["A", "B", "C"]:
        conf = f"Mode{mode}_Confirmed"
        confirmed_entries = df[df[conf] == True]
        
        for idx in confirmed_entries.index:
            row = df.loc[idx]
            
            # Get mode-specific parameters
            orb_low = row.get("ORBLow", None) if mode == "A" else None
            
            # For Mode C, calculate consolidation low (last 3-6 candles before entry)
            consolidation_low = None
            if mode == "C":
                # Get previous 3-6 candles from same day
                same_day = df[(df["Symbol"] == row["Symbol"]) & 
                             (df["Date"] == row["Date"]) &
                             (df.index < idx)]
                if len(same_day) >= 3:
                    recent_candles = same_day.tail(6)
                    consolidation_low = recent_candles["Low"].min()
            
            # Calculate stop-loss and target
            sl_data = calculate_stop_loss_and_target(
                row, 
                mode, 
                orb_low=orb_low,
                consolidation_low=consolidation_low
            )
            
            if sl_data:
                df.loc[idx, f"Mode{mode}_Stop_ATR"] = sl_data["Stop_ATR"]
                df.loc[idx, f"Mode{mode}_Stop_Structure"] = sl_data["Stop_Structure"]
                df.loc[idx, f"Mode{mode}_Final_Stop"] = sl_data["Final_Stop"]
                df.loc[idx, f"Mode{mode}_Target"] = sl_data["Target"]
                df.loc[idx, f"Mode{mode}_Risk_Per_Share"] = sl_data["Risk_Per_Share"]
                df.loc[idx, f"Mode{mode}_Risk_Pct"] = sl_data["Risk_Pct"]

    return df

# ======================================================
# RUN
# ======================================================

if __name__ == "__main__":
    print("Loading Phase 2 results...")
    phase2_df = load_phase2()
    print(f"Found {len(phase2_df)} Phase-2 qualified entries ({phase2_df['Symbol'].nunique()} unique symbols)")
    
    print("\nLoading NSEI data...")
    nsei_df = load_nsei_5m()
    print(f"Loaded {len(nsei_df)} NSEI 5m candles")

    results = []
    symbols = phase2_df["Symbol"].unique()
    
    print(f"\nProcessing {len(symbols)} symbols...")
    for i, symbol in enumerate(symbols, 1):
        print(f"  [{i}/{len(symbols)}] Processing {symbol}...", end="\r")
        out = run_phase3_for_symbol(symbol, phase2_df, nsei_df)
        if out is not None:
            results.append(out)

    if not results:
        print("\n\n❌ No results generated. Check if Phase 2 data and stock 5m data are aligned.")
        exit(1)

    final_df = pd.concat(results, ignore_index=True)
    
    # Convert timezone-aware datetime to naive for Excel compatibility
    if 'Datetime' in final_df.columns:
        final_df['Datetime'] = final_df['Datetime'].dt.tz_localize(None)
    
    # Save to Excel with simple table format
    output_file = "phase-3results/Phase_3_Entrieesss.xlsx"
    os.makedirs("phase-3results", exist_ok=True)
    
    # Extract confirmed entries for each mode
    mode_a = final_df[final_df["ModeA_Confirmed"]].copy()
    mode_b = final_df[final_df["ModeB_Confirmed"]].copy()
    mode_c = final_df[final_df["ModeC_Confirmed"]].copy()
    
    # Prepare unified table
    entries = []
    
    # Add Mode A entries with clear descriptions
    for _, row in mode_a.iterrows():
        # Get trigger from this row (Mode A always has trigger on confirmed candle)
        trigger_val = round(row["ModeA_Trigger"], 2) if pd.notna(row["ModeA_Trigger"]) else None
        
        entries.append({
            "Date": row["Date"],
            "Stock": row["Symbol"],
            "Entry Mode": "A - ORB Breakout",
            "Entry Time": row["Time"],
            "Entry Price (₹)": round(row["ModeA_Entry"], 2),
            "Stop-Loss (₹)": row.get("ModeA_Final_Stop", "—"),
            "Target (₹)": row.get("ModeA_Target", "—"),
            "Risk Per Share (₹)": row.get("ModeA_Risk_Per_Share", "—"),
            "Risk %": row.get("ModeA_Risk_Pct", "—"),
            "Trigger Was (₹)": trigger_val,
            "Volume Strength": f"{round(row['VolMult_od'], 2)}x",
            "Relative Strength %": round(row["RS_30m"], 2) if pd.notna(row["RS_30m"]) else "—",
            "Why Entered": "Price broke above Opening Range High with strong volume and momentum"
        })
    
    # Add Mode B entries with clear descriptions
    for _, row in mode_b.iterrows():
        # Get trigger from first eligible candle for this symbol-date
        trigger_val = None
        if pd.notna(row["ModeB_Trigger"]):
            trigger_val = round(row["ModeB_Trigger"], 2)
        else:
            # Find first candle with ModeB_Trigger set for this symbol-date
            same_day = final_df[(final_df["Symbol"] == row["Symbol"]) & 
                               (final_df["Date"] == row["Date"]) & 
                               (final_df["ModeB_Trigger"].notna())]
            if len(same_day) > 0:
                trigger_val = round(same_day.iloc[0]["ModeB_Trigger"], 2)
        
        entries.append({
            "Date": row["Date"],
            "Stock": row["Symbol"],
            "Entry Mode": "B - VWAP Reclaim",
            "Entry Time": row["Time"],
            "Entry Price (₹)": round(row["ModeB_Entry"], 2),
            "Stop-Loss (₹)": row.get("ModeB_Final_Stop", "—"),
            "Target (₹)": row.get("ModeB_Target", "—"),
            "Risk Per Share (₹)": row.get("ModeB_Risk_Per_Share", "—"),
            "Risk %": row.get("ModeB_Risk_Pct", "—"),
            "Trigger Was (₹)": trigger_val,
            "Volume Strength": f"{round(row['VolMult_od'], 2)}x",
            "Relative Strength %": round(row["RS_30m"], 2) if pd.notna(row["RS_30m"]) else "—",
            "Why Entered": "Price pulled back to VWAP and reclaimed it with volume support"
        })
    
    # Add Mode C entries with clear descriptions
    for _, row in mode_c.iterrows():
        # Get trigger from first eligible candle for this symbol-date
        trigger_val = None
        if pd.notna(row["ModeC_Trigger"]):
            trigger_val = round(row["ModeC_Trigger"], 2)
        else:
            # Find first candle with ModeC_Trigger set for this symbol-date
            same_day = final_df[(final_df["Symbol"] == row["Symbol"]) & 
                               (final_df["Date"] == row["Date"]) & 
                               (final_df["ModeC_Trigger"].notna())]
            if len(same_day) > 0:
                trigger_val = round(same_day.iloc[0]["ModeC_Trigger"], 2)
        
        entries.append({
            "Date": row["Date"],
            "Stock": row["Symbol"],
            "Entry Mode": "C - Day High Break",
            "Entry Time": row["Time"],
            "Entry Price (₹)": round(row["ModeC_Entry"], 2),
            "Stop-Loss (₹)": row.get("ModeC_Final_Stop", "—"),
            "Target (₹)": row.get("ModeC_Target", "—"),
            "Risk Per Share (₹)": row.get("ModeC_Risk_Per_Share", "—"),
            "Risk %": row.get("ModeC_Risk_Pct", "—"),
            "Trigger Was (₹)": trigger_val,
            "Volume Strength": f"{round(row['VolMult_od'], 2)}x",
            "Relative Strength %": round(row["RS_30m"], 2) if pd.notna(row["RS_30m"]) else "—",
            "Why Entered": "Price consolidated near day high and broke out with strong volume"
        })
    
    # Create DataFrame with user-friendly columns
    if entries:
        entries_df = pd.DataFrame(entries)
        # Sort by date and time
        entries_df = entries_df.sort_values(["Date", "Entry Time"]).reset_index(drop=True)
    else:
        # Empty DataFrame with columns
        entries_df = pd.DataFrame(columns=[
            "Date", "Stock", "Entry Mode", "Entry Time", "Entry Price (₹)", 
            "Stop-Loss (₹)", "Target (₹)", "Risk Per Share (₹)", "Risk %",
            "Trigger Was (₹)", "Volume Strength", "Relative Strength %", "Why Entered"
        ])
    
    # Create Excel with formatting
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Entry Signals"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, col_name in enumerate(entries_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_idx, row_data in enumerate(dataframe_to_rows(entries_df, index=False, header=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Color code by mode
            if col_idx == 3:  # Mode column
                if value == "A":
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                elif value == "B":
                    cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                elif value == "C":
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 14  # Stock
    ws.column_dimensions['C'].width = 18  # Entry Mode
    ws.column_dimensions['D'].width = 12  # Entry Time
    ws.column_dimensions['E'].width = 14  # Entry Price
    ws.column_dimensions['F'].width = 14  # Stop-Loss
    ws.column_dimensions['G'].width = 14  # Target
    ws.column_dimensions['H'].width = 16  # Risk Per Share
    ws.column_dimensions['I'].width = 10  # Risk %
    ws.column_dimensions['J'].width = 14  # Trigger Was
    ws.column_dimensions['K'].width = 16  # Volume Strength
    ws.column_dimensions['L'].width = 18  # Relative Strength
    ws.column_dimensions['M'].width = 60  # Why Entered (wide for explanation)
    
    
    # Add All Data sheet - USER FRIENDLY VERSION with plain English
    ws2 = wb.create_sheet("All Data (Reference)")
    
    # Select only important columns and rename them to plain English
    all_data_display = final_df[[
        "Symbol", "Date", "Time", "Open", "High", "Low", "Close", "Volume",
        "VWAP", "VolMult_od", "RS_30m", "ATR_pct", "ATR_5m_pct", "Buffer",
        "ModeA_Eligible", "ModeB_Eligible", "ModeC_Eligible",
        "ModeA_Trigger", "ModeB_Trigger", "ModeC_Trigger",
        "ModeA_Confirmed", "ModeB_Confirmed", "ModeC_Confirmed",
        "ModeA_Final_Stop", "ModeB_Final_Stop", "ModeC_Final_Stop",
        "ModeA_Target", "ModeB_Target", "ModeC_Target"
    ]].copy()
    
    # Rename columns to plain English
    all_data_display.columns = [
        "Stock", "Date", "Time", "Open Price", "High Price", "Low Price", "Close Price", "Volume",
        "VWAP (Avg Price)", "Volume Strength", "Relative Strength %", "Daily Volatility %", "Intraday ATR %", "Buffer Amount",
        "Mode A Eligible?", "Mode B Eligible?", "Mode C Eligible?",
        "Mode A Trigger", "Mode B Trigger", "Mode C Trigger",
        "Mode A Confirmed?", "Mode B Confirmed?", "Mode C Confirmed?",
        "Mode A Stop-Loss", "Mode B Stop-Loss", "Mode C Stop-Loss",
        "Mode A Target", "Mode B Target", "Mode C Target"
    ]
    
    # Write to sheet with formatting
    for r_idx, row in enumerate(dataframe_to_rows(all_data_display, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    ws2.column_dimensions['A'].width = 14  # Stock
    ws2.column_dimensions['B'].width = 12  # Date
    ws2.column_dimensions['C'].width = 10  # Time
    ws2.column_dimensions['D'].width = 12  # Open
    ws2.column_dimensions['E'].width = 12  # High
    ws2.column_dimensions['F'].width = 12  # Low
    ws2.column_dimensions['G'].width = 12  # Close
    ws2.column_dimensions['H'].width = 12  # Volume
    ws2.column_dimensions['I'].width = 14  # VWAP
    ws2.column_dimensions['J'].width = 16  # Volume Strength
    ws2.column_dimensions['K'].width = 18  # Relative Strength
    ws2.column_dimensions['L'].width = 12  # Volatility
    ws2.column_dimensions['M'].width = 14  # Buffer
    ws2.column_dimensions['N'].width = 16  # Mode A Eligible
    ws2.column_dimensions['O'].width = 16  # Mode B Eligible
    ws2.column_dimensions['P'].width = 16  # Mode C Eligible
    ws2.column_dimensions['Q'].width = 14  # Mode A Trigger
    ws2.column_dimensions['R'].width = 14  # Mode B Trigger
    ws2.column_dimensions['S'].width = 14  # Mode C Trigger
    ws2.column_dimensions['T'].width = 18  # Mode A Confirmed
    ws2.column_dimensions['U'].width = 18  # Mode B Confirmed
    ws2.column_dimensions['V'].width = 18  # Mode C Confirmed
    
    
    # ================= MODE-SPECIFIC SHEETS (279 ROWS EACH) =================
    # Add 3 separate sheets for each mode showing ONE row per stock-date
    
    # Helper function to get one representative row per stock-date
    def get_representative_row(df, mode):
        """
        For each stock-date, pick ONE row to display:
        Priority: Confirmed > Eligible > First candle of day
        """
        result_rows = []
        
        for (symbol, date), group in df.groupby(['Symbol', 'Date']):
            group = group.sort_values('Datetime')
            
            # Try to find confirmed row first
            confirmed = group[group[f'Mode{mode}_Confirmed'] == True]
            if not confirmed.empty:
                result_rows.append(confirmed.iloc[0])
                continue
            
            # Try to find eligible row
            eligible = group[group[f'Mode{mode}_Eligible'] == True]
            if not eligible.empty:
                result_rows.append(eligible.iloc[0])
                continue
            
            # Otherwise, take first candle of the day (representative)
            result_rows.append(group.iloc[0])
        
        return pd.DataFrame(result_rows)
    
    # Sheet 3: Mode A - ORB Breakout (279 rows)
    ws3 = wb.create_sheet("Mode A - ORB Breakout")
    mode_a_summary = get_representative_row(final_df, 'A')
    
    mode_a_data = mode_a_summary[[
        "Date", "Symbol", "Time", "Open", "High", "Low", "Close", "Volume",
        "VWAP", "VolMult_od", "RS_30m", "ORBHigh", "ORBLow",
        "ModeA_Eligible", "ModeA_Trigger", "ModeA_Confirmed", "ModeA_Entry",
        "ModeA_Final_Stop", "ModeA_Target", "ModeA_Risk_Pct"
    ]].copy()
    
    mode_a_data.columns = [
        "Date", "Stock", "Time", "Open", "High", "Low", "Close", "Volume",
        "VWAP", "Vol Strength", "RS %", "ORB High", "ORB Low",
        "Eligible?", "Trigger", "Confirmed?", "Entry Price",
        "Stop-Loss", "Target", "Risk %"
    ]
    
    # Write Mode A data
    for r_idx, row in enumerate(dataframe_to_rows(mode_a_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")
                # Highlight confirmed rows in green, eligible in yellow
                if r_idx > 1:  # Skip header
                    row_data = list(mode_a_data.iloc[r_idx-2])
                    is_confirmed = row_data[15]  # Confirmed? column
                    is_eligible = row_data[13]   # Eligible? column
                    
                    if is_confirmed:
                        for col in range(1, 21):
                            ws3.cell(row=r_idx, column=col).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    elif is_eligible:
                        for col in range(1, 21):
                            ws3.cell(row=r_idx, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Set column widths for Mode A
    for col, width in zip(['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T'],
                          [12,14,10,10,10,10,10,12,12,12,10,12,12,12,12,12,12,12,12,10]):
        ws3.column_dimensions[col].width = width
    
    # Sheet 4: Mode B - VWAP Reclaim (279 rows)
    ws4 = wb.create_sheet("Mode B - VWAP Reclaim")
    mode_b_summary = get_representative_row(final_df, 'B')
    
    mode_b_data = mode_b_summary[[
        "Date", "Symbol", "Time", "Open", "High", "Low", "Close", "Volume",
        "VWAP", "VolMult_od", "RS_30m", "NearVWAP",
        "ModeB_Eligible", "ModeB_Trigger", "ModeB_Confirmed", "ModeB_Entry",
        "ModeB_Final_Stop", "ModeB_Target", "ModeB_Risk_Pct"
    ]].copy()
    
    mode_b_data.columns = [
        "Date", "Stock", "Time", "Open", "High", "Low", "Close", "Volume",
        "VWAP", "Vol Strength", "RS %", "Near VWAP?",
        "Eligible?", "Trigger", "Confirmed?", "Entry Price",
        "Stop-Loss", "Target", "Risk %"
    ]
    
    # Write Mode B data
    for r_idx, row in enumerate(dataframe_to_rows(mode_b_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")
                # Highlight confirmed rows in blue, eligible in yellow
                if r_idx > 1:  # Skip header
                    row_data = list(mode_b_data.iloc[r_idx-2])
                    is_confirmed = row_data[14]  # Confirmed? column
                    is_eligible = row_data[12]   # Eligible? column
                    
                    if is_confirmed:
                        for col in range(1, 20):
                            ws4.cell(row=r_idx, column=col).fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                    elif is_eligible:
                        for col in range(1, 20):
                            ws4.cell(row=r_idx, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Set column widths for Mode B
    for col, width in zip(['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S'],
                          [12,14,10,10,10,10,10,12,12,12,10,12,12,12,12,12,12,12,10]):
        ws4.column_dimensions[col].width = width
    
    # Sheet 5: Mode C - Day High Break (279 rows)
    ws5 = wb.create_sheet("Mode C - Day High Break")
    mode_c_summary = get_representative_row(final_df, 'C')
    
    mode_c_data = mode_c_summary[[
        "Date", "Symbol", "Time", "Open", "High", "Low", "Close", "Volume",
        "VWAP", "VolMult_od", "DayHigh", "NearHigh",
        "ModeC_Eligible", "ModeC_Trigger", "ModeC_Confirmed", "ModeC_Entry",
        "ModeC_Final_Stop", "ModeC_Target", "ModeC_Risk_Pct"
    ]].copy()
    
    mode_c_data.columns = [
        "Date", "Stock", "Time", "Open", "High", "Low", "Close", "Volume",
        "VWAP", "Vol Strength", "Day High", "Near High?",
        "Eligible?", "Trigger", "Confirmed?", "Entry Price",
        "Stop-Loss", "Target", "Risk %"
    ]
    
    # Write Mode C data
    for r_idx, row in enumerate(dataframe_to_rows(mode_c_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")
                # Highlight confirmed rows in orange, eligible in yellow
                if r_idx > 1:  # Skip header
                    row_data = list(mode_c_data.iloc[r_idx-2])
                    is_confirmed = row_data[14]  # Confirmed? column
                    is_eligible = row_data[12]   # Eligible? column
                    
                    if is_confirmed:
                        for col in range(1, 20):
                            ws5.cell(row=r_idx, column=col).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                    elif is_eligible:
                        for col in range(1, 20):
                            ws5.cell(row=r_idx, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Set column widths for Mode C
    for col, width in zip(['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S'],
                          [12,14,10,10,10,10,10,12,12,12,12,12,12,12,12,12,12,12,10]):
        ws5.column_dimensions[col].width = width
    
    # Save workbook
    wb.save(output_file)
    
    
    print(f"\n\n✅ Results saved to: {output_file}")
    
    # Summary
    print("\n" + "="*60)
    print("PHASE 3 - ENTRY CONFIRMATION SUMMARY")
    print("="*60)
    
    mode_a_count = final_df["ModeA_Confirmed"].sum()
    mode_b_count = final_df["ModeB_Confirmed"].sum()
    mode_c_count = final_df["ModeC_Confirmed"].sum()
    
    print(f"\nMODE A (ORB Breakout): {mode_a_count} confirmed entries")
    if mode_a_count > 0:
        print(final_df[final_df["ModeA_Confirmed"]][
            ["Symbol", "Date", "Time", "ModeA_Trigger", "ModeA_Entry"]
        ].to_string(index=False))
    
    print(f"\nMODE B (VWAP Reclaim): {mode_b_count} confirmed entries")
    if mode_b_count > 0:
        print(final_df[final_df["ModeB_Confirmed"]][
            ["Symbol", "Date", "Time", "ModeB_Trigger", "ModeB_Entry"]
        ].to_string(index=False))
    
    print(f"\nMODE C (Day-High Continuation): {mode_c_count} confirmed entries")
    if mode_c_count > 0:
        print(final_df[final_df["ModeC_Confirmed"]][
            ["Symbol", "Date", "Time", "ModeC_Trigger", "ModeC_Entry"]
        ].to_string(index=False))
    
    print(f"\n{'='*60}")
    print(f"TOTAL CONFIRMED ENTRIES: {mode_a_count + mode_b_count + mode_c_count}")
    print(f"{'='*60}\n")

